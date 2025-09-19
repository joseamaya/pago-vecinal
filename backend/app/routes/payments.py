from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form
from fastapi.responses import FileResponse
from typing import List, Optional
from pydantic import BaseModel
from datetime import datetime
import os
import shutil
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from ..models.payment import Payment, PaymentCreate, PaymentUpdate, PaymentResponse, PaymentStatus
from ..models.fee import Fee, FeeStatus
from ..models.user import User, UserRole
from ..models.receipt import Receipt
from ..models.property import Property
from ..routes.auth import get_current_user

class BulkApproveRequest(BaseModel):
    payment_ids: List[str]

class PaginatedPaymentResponse(BaseModel):
    data: List[PaymentResponse]
    pagination: dict

router = APIRouter()

@router.get("/", response_model=PaginatedPaymentResponse)
async def get_payments(page: int = 1, limit: int = 20, current_user: User = Depends(get_current_user)):
    # Build query filters
    query_filters = {}
    if current_user.role != UserRole.ADMIN:
        # Owners can only see their own payments
        query_filters["user.id"] = current_user.id

    # Get total count
    if query_filters:
        total_count = await Payment.find(query_filters).count()
    else:
        total_count = await Payment.count()

    # Calculate skip
    skip = (page - 1) * limit

    # Get paginated payments, sorted by payment_date descending
    if query_filters:
        payments = await Payment.find(query_filters).sort([("payment_date", -1)]).skip(skip).limit(limit).to_list()
    else:
        payments = await Payment.find_all().sort([("payment_date", -1)]).skip(skip).limit(limit).to_list()

    # Fetch links for each payment
    for payment in payments:
        await payment.fetch_link(Payment.fee)
        await payment.fetch_link(Payment.user)

    # Calculate total pages
    total_pages = (total_count + limit - 1) // limit

    payment_responses = [
        PaymentResponse(
            id=str(payment.id),
            fee_id=payment.fee_id,
            user_id=str(payment.user.id),
            amount=payment.amount,
            payment_date=payment.payment_date,
            receipt_file=payment.receipt_file,
            generated_receipt_file=payment.generated_receipt_file,
            status=payment.status,
            notes=payment.notes
        )
        for payment in payments
    ]

    return PaginatedPaymentResponse(
        data=payment_responses,
        pagination={
            "page": page,
            "limit": limit,
            "total_count": total_count,
            "total_pages": total_pages
        }
    )

@router.get("/{payment_id}", response_model=PaymentResponse)
async def get_payment(payment_id: str, current_user: User = Depends(get_current_user)):
    payment = await Payment.get(payment_id)
    if not payment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Payment not found"
        )

    # Fetch linked documents
    await payment.fetch_link(Payment.fee)
    await payment.fetch_link(Payment.user)

    # Check permissions
    if (current_user.role != UserRole.ADMIN and
        str(payment.user.id) != str(current_user.id)):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions"
        )

    return PaymentResponse(
        id=str(payment.id),
        fee_id=payment.fee_id,
        user_id=str(payment.user.id),
        amount=payment.amount,
        payment_date=payment.payment_date,
        receipt_file=payment.receipt_file,
        generated_receipt_file=payment.generated_receipt_file,
        status=payment.status,
        notes=payment.notes
    )

@router.post("/", response_model=PaymentResponse)
async def create_payment(
    fee_id: str = Form(...),
    amount: float = Form(...),
    notes: Optional[str] = Form(None),
    receipt_file: Optional[UploadFile] = File(None),
    current_user: User = Depends(get_current_user)
):
    # Get the fee
    fee = await Fee.get(fee_id)
    if not fee:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Fee not found"
        )

    # Check if user owns this fee or is admin
    if (current_user.role != UserRole.ADMIN and
        (not fee.user or str(fee.user.id) != str(current_user.id))):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions"
        )

    # Handle file upload
    receipt_file_path = None
    if receipt_file:
        # Create unique filename
        file_extension = os.path.splitext(receipt_file.filename)[1]
        unique_filename = f"{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{current_user.id}{file_extension}"
        file_path = os.path.join("static", "uploads", unique_filename)

        # Save file
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(receipt_file.file, buffer)

        receipt_file_path = file_path

    payment = Payment(
        fee=fee,
        fee_id=str(fee.id),
        user=current_user,
        amount=amount,
        payment_date=datetime.utcnow(),
        receipt_file=receipt_file_path,
        notes=notes
    )
    await payment.insert()

    # Fetch linked documents for the response
    await payment.fetch_link(Payment.fee)
    await payment.fetch_link(Payment.user)

    return PaymentResponse(
        id=str(payment.id),
        fee_id=payment.fee_id,
        user_id=str(payment.user.id),
        amount=payment.amount,
        payment_date=payment.payment_date,
        receipt_file=payment.receipt_file,
        generated_receipt_file=payment.generated_receipt_file,
        status=payment.status,
        notes=payment.notes
    )

@router.put("/{payment_id}", response_model=PaymentResponse)
async def update_payment(
    payment_id: str,
    amount: Optional[float] = Form(None),
    status: Optional[str] = Form(None),
    notes: Optional[str] = Form(None),
    receipt_file: Optional[UploadFile] = File(None),
    current_user: User = Depends(get_current_user)
):
    payment = await Payment.get(payment_id)
    if not payment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Payment not found"
        )

    # Fetch linked documents
    await payment.fetch_link(Payment.fee)
    await payment.fetch_link(Payment.user)

    # Check permissions
    if (current_user.role != UserRole.ADMIN and
        str(payment.user.id) != str(current_user.id)):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions"
        )

    # Handle file upload
    if receipt_file:
        # Create unique filename
        file_extension = os.path.splitext(receipt_file.filename)[1]
        unique_filename = f"{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{current_user.id}{file_extension}"
        file_path = os.path.join("static", "uploads", unique_filename)

        # Save file
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(receipt_file.file, buffer)

        payment.receipt_file = file_path

    # Update fields
    if amount is not None:
        payment.amount = amount
    if status is not None:
        payment.status = PaymentStatus(status)
    if notes is not None:
        payment.notes = notes

    # Create receipt in database if status changed to approved
    if status == "approved":
        try:
            print(f"Creating receipt for payment {payment_id}")

            # Fetch property and fee_schedule from fee if not already fetched
            if payment.fee:
                await payment.fee.fetch_link('property')
                await payment.fee.fetch_link('fee_schedule')

            # Generate correlative number
            current_year = datetime.utcnow().year
            last_receipt = await Receipt.find(
                {"correlative_number": {"$regex": f"^REC-{current_year}"}}
            ).sort([("correlative_number", -1)]).first_or_none()

            if last_receipt:
                parts = last_receipt.correlative_number.split("-")
                if len(parts) == 3:
                    last_number = int(parts[2])
                    new_number = last_number + 1
                else:
                    new_number = 1
            else:
                new_number = 1

            correlative_number = f"REC-{current_year}-{new_number:05d}"

            # Create property and owner details snapshot
            if not payment.fee or not payment.fee.property:
                # If no property linked, create default details
                property_details = {
                    "villa": "N/A",
                    "row_letter": "N/A",
                    "number": 0,
                    "owner_name": "Propietario no registrado",
                    "owner_phone": "N/A"
                }
                owner_details = {
                    "name": "Propietario no registrado",
                    "phone": "N/A"
                }
            else:
                property_details = {
                    "villa": getattr(payment.fee.property, 'villa', 'N/A'),
                    "row_letter": getattr(payment.fee.property, 'row_letter', 'N/A'),
                    "number": getattr(payment.fee.property, 'number', 0),
                    "owner_name": getattr(payment.fee.property, 'owner_name', 'Propietario no registrado'),
                    "owner_phone": getattr(payment.fee.property, 'owner_phone', 'N/A') or "N/A"
                }
                owner_details = {
                    "name": getattr(payment.fee.property, 'owner_name', 'Propietario no registrado'),
                    "phone": getattr(payment.fee.property, 'owner_phone', 'N/A') or "N/A"
                }

            # Create receipt record
            receipt = Receipt(
                correlative_number=correlative_number,
                payment=payment,
                issue_date=datetime.utcnow(),
                total_amount=payment.amount,
                property_details=property_details,
                owner_details=owner_details,
                fee_period=f"Cuota {payment.fee.reference or 'N/A'}" if payment.fee else "N/A",
                notes=f"Recibo generado automáticamente al aprobar el pago"
            )

            await receipt.insert()

            print(f"Receipt created in database with ID: {receipt.id}")
        except Exception as e:
            # Log the error but don't fail the payment update
            print(f"Error creating receipt for payment {payment_id}: {e}")
            import traceback
            traceback.print_exc()

    await payment.save()

    # Update fee status to completed if payment was approved
    if status == "approved" and payment.fee:
        payment.fee.status = FeeStatus.COMPLETED
        await payment.fee.save()

    print(f"Payment saved with generated_receipt_file: {payment.generated_receipt_file}")

    # Fetch links again after save
    await payment.fetch_link(Payment.fee)
    await payment.fetch_link(Payment.user)

    response_data = PaymentResponse(
        id=str(payment.id),
        fee_id=payment.fee_id,
        user_id=str(payment.user.id),
        amount=payment.amount,
        payment_date=payment.payment_date,
        receipt_file=payment.receipt_file,
        generated_receipt_file=payment.generated_receipt_file,
        status=payment.status,
        notes=payment.notes
    )
    print(f"Returning response with generated_receipt_file: {response_data.generated_receipt_file}")
    return response_data

@router.delete("/{payment_id}")
async def delete_payment(payment_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions"
        )

    payment = await Payment.get(payment_id)
    if not payment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Payment not found"
        )
    await payment.delete()
    return {"message": "Payment deleted successfully"}

@router.get("/{payment_id}/download-receipt")
async def download_generated_receipt(payment_id: str, current_user: User = Depends(get_current_user)):
    """Download the automatically generated receipt PDF for a payment"""
    payment = await Payment.get(payment_id)
    if not payment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Payment not found"
        )

    # Fetch linked documents
    await payment.fetch_link(Payment.fee)
    await payment.fetch_link(Payment.user)

    # Check permissions
    if (current_user.role != UserRole.ADMIN and
        str(payment.user.id) != str(current_user.id)):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions"
        )

    # Check if receipt file exists
    if not payment.generated_receipt_file:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No receipt file associated with this payment"
        )

    # Convert relative path to absolute path if needed
    file_path = payment.generated_receipt_file
    if not os.path.isabs(file_path):
        file_path = os.path.join(os.getcwd(), file_path)

    print(f"Looking for receipt file at: {file_path}")
    print(f"File exists: {os.path.exists(file_path)}")

    if not os.path.exists(file_path):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Receipt file not found at {file_path}"
        )

    # Return the PDF file
    filename = f"recibo_pago_{payment_id}.pdf"
    return FileResponse(
        path=file_path,
        filename=filename,
        media_type="application/pdf"
    )

@router.post("/bulk-import")
async def bulk_import_payments(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """Bulk import payments from Excel file"""
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions"
        )

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File must be an Excel file (.xlsx or .xls)"
        )

    try:
        # Load workbook
        wb = load_workbook(file.file)
        ws = wb.active

        # Expected columns: Villa, Fila, Número, Año, Mes, Monto, Fecha de Pago, Notas
        expected_headers = ['Villa', 'Fila', 'Número', 'Año', 'Mes', 'Monto', 'Fecha de Pago', 'Notas']

        # Check headers
        actual_headers = []
        for col in range(1, 9):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                actual_headers.append(str(cell_value).strip())
            else:
                actual_headers.append("")

        if actual_headers[:7] != expected_headers[:7]:  # Notas is optional
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid Excel format. Expected headers: {', '.join(expected_headers[:7])}"
            )

        results = {
            "successful_imports": 0,
            "failed_imports": 0,
            "errors": []
        }

        # Process each row starting from row 2
        for row_idx in range(2, ws.max_row + 1):
            try:
                # Extract data
                villa = str(ws.cell(row=row_idx, column=1).value or "").strip()
                fila = str(ws.cell(row=row_idx, column=2).value or "").strip()
                numero = ws.cell(row=row_idx, column=3).value
                year = ws.cell(row=row_idx, column=4).value
                month = ws.cell(row=row_idx, column=5).value
                amount = ws.cell(row=row_idx, column=6).value
                payment_date_cell = ws.cell(row=row_idx, column=7).value
                notes = str(ws.cell(row=row_idx, column=8).value or "").strip()

                # Validate required fields
                if not all([villa, fila, numero is not None, year is not None, month is not None, amount is not None, payment_date_cell is not None]):
                    results["errors"].append({
                        "row": row_idx,
                        "error": "Missing required fields"
                    })
                    results["failed_imports"] += 1
                    continue

                # Convert types
                try:
                    numero = int(numero)
                    year = int(year)
                    month = int(month)
                    amount = float(amount)
                    print(payment_date_cell)

                    if isinstance(payment_date_cell, datetime):
                        payment_date = payment_date_cell
                    else:
                        payment_date = datetime.strptime(str(payment_date_cell), '%Y-%m-%d')

                    print(f"Parsed payment_date: {payment_date}")

                except (ValueError, TypeError) as e:
                    results["errors"].append({
                        "row": row_idx,
                        "error": f"Invalid data types: {str(e)}"
                    })
                    results["failed_imports"] += 1
                    continue

                # Validate data ranges
                if month < 1 or month > 12:
                    results["errors"].append({
                        "row": row_idx,
                        "error": "Month must be between 1 and 12"
                    })
                    results["failed_imports"] += 1
                    continue

                if amount <= 0:
                    results["errors"].append({
                        "row": row_idx,
                        "error": "Amount must be greater than 0"
                    })
                    results["failed_imports"] += 1
                    continue

                if payment_date > datetime.utcnow():
                    results["errors"].append({
                        "row": row_idx,
                        "error": "Payment date cannot be in the future"
                    })
                    results["failed_imports"] += 1
                    continue

                # Find property
                property_obj = await Property.find_one(
                    Property.villa == villa,
                    Property.row_letter == fila,
                    Property.number == numero
                )

                if not property_obj:
                    results["errors"].append({
                        "row": row_idx,
                        "error": f"Property not found: Villa {villa}, {fila}{numero}"
                    })
                    results["failed_imports"] += 1
                    continue

                # Find fee for this property and period
                fee = await Fee.find_one(
                    Fee.property.id == property_obj.id,
                    Fee.year == year,
                    Fee.month == month,
                    Fee.status == FeeStatus.PENDING
                )

                if not fee:
                    results["errors"].append({
                        "row": row_idx,
                        "error": f"No pending fee found for property {villa}-{fila}{numero} in {month}/{year}"
                    })
                    results["failed_imports"] += 1
                    continue

                # Check if payment already exists for this fee
                existing_payment = await Payment.find_one(Payment.fee_id == str(fee.id))
                if existing_payment:
                    results["errors"].append({
                        "row": row_idx,
                        "error": f"Payment already exists for this fee"
                    })
                    results["failed_imports"] += 1
                    continue

                # Create payment
                payment = Payment(
                    fee=fee,
                    fee_id=str(fee.id),
                    user=current_user,  # Admin user creating the bulk import
                    amount=amount,
                    payment_date=payment_date,
                    notes=notes if notes else None
                )
                await payment.insert()

                # Update fee status to completed
                fee.status = FeeStatus.COMPLETED
                await fee.save()

                results["successful_imports"] += 1

            except Exception as e:
                results["errors"].append({
                    "row": row_idx,
                    "error": f"Unexpected error: {str(e)}"
                })
                results["failed_imports"] += 1
                continue

        return results

    except InvalidFileException:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid Excel file"
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error processing file: {str(e)}"
        )

@router.post("/bulk-approve")
async def bulk_approve_payments(
    request: BulkApproveRequest,
    current_user: User = Depends(get_current_user)
):
    payment_ids = request.payment_ids
    """Bulk approve multiple payments"""
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions"
        )

    if not payment_ids:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No payment IDs provided"
        )

    approved_count = 0
    errors = []

    for payment_id in payment_ids:
        try:
            payment = await Payment.get(payment_id)
            if not payment:
                errors.append(f"Payment {payment_id} not found")
                continue

            # Check if payment is already approved
            if payment.status == PaymentStatus.APPROVED:
                errors.append(f"Payment {payment_id} is already approved")
                continue

            # Fetch linked documents
            await payment.fetch_link(Payment.fee)
            await payment.fetch_link(Payment.user)

            # Update payment status
            payment.status = PaymentStatus.APPROVED
            await payment.save()

            # Update fee status to completed
            if payment.fee:
                payment.fee.status = FeeStatus.COMPLETED
                await payment.fee.save()

            # Generate receipt (similar to individual approval)
            try:
                print(f"Creating receipt for payment {payment_id}")

                # Fetch property and fee_schedule from fee if not already fetched
                if payment.fee:
                    await payment.fee.fetch_link('property')
                    await payment.fee.fetch_link('fee_schedule')

                # Generate correlative number
                current_year = datetime.utcnow().year
                last_receipt = await Receipt.find(
                    {"correlative_number": {"$regex": f"^REC-{current_year}"}}
                ).sort([("correlative_number", -1)]).first_or_none()

                if last_receipt:
                    parts = last_receipt.correlative_number.split("-")
                    if len(parts) == 3:
                        last_number = int(parts[2])
                        new_number = last_number + 1
                    else:
                        new_number = 1
                else:
                    new_number = 1

                correlative_number = f"REC-{current_year}-{new_number:05d}"

                # Create property and owner details snapshot
                if not payment.fee or not payment.fee.property:
                    # If no property linked, create default details
                    property_details = {
                        "villa": "N/A",
                        "row_letter": "N/A",
                        "number": 0,
                        "owner_name": "Propietario no registrado",
                        "owner_phone": "N/A"
                    }
                    owner_details = {
                        "name": "Propietario no registrado",
                        "phone": "N/A"
                    }
                else:
                    property_details = {
                        "villa": getattr(payment.fee.property, 'villa', 'N/A'),
                        "row_letter": getattr(payment.fee.property, 'row_letter', 'N/A'),
                        "number": getattr(payment.fee.property, 'number', 0),
                        "owner_name": getattr(payment.fee.property, 'owner_name', 'Propietario no registrado'),
                        "owner_phone": getattr(payment.fee.property, 'owner_phone', 'N/A') or "N/A"
                    }
                    owner_details = {
                        "name": getattr(payment.fee.property, 'owner_name', 'Propietario no registrado'),
                        "phone": getattr(payment.fee.property, 'owner_phone', 'N/A') or "N/A"
                    }

                # Create receipt record
                receipt = Receipt(
                    correlative_number=correlative_number,
                    payment=payment,
                    issue_date=datetime.utcnow(),
                    total_amount=payment.amount,
                    property_details=property_details,
                    owner_details=owner_details,
                    fee_period=f"Cuota {payment.fee.reference or 'N/A'}" if payment.fee else "N/A",
                    notes=f"Recibo generado automáticamente al aprobar el pago (aprobación masiva)"
                )

                await receipt.insert()

                print(f"Receipt created in database with ID: {receipt.id}")
            except Exception as e:
                # Log the error but don't fail the bulk operation
                print(f"Error creating receipt for payment {payment_id}: {e}")
                import traceback
                traceback.print_exc()

            approved_count += 1

        except Exception as e:
            errors.append(f"Error processing payment {payment_id}: {str(e)}")

    return {
        "message": f"Successfully approved {approved_count} payments",
        "approved_count": approved_count,
        "errors": errors
    }