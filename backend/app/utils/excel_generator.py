import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_property_payment_history_excel(property_data, payments_data, fees_data):
    """
    Generate an Excel report of payment history for a specific property
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Historial de Pagos"

    # Title
    ws['A1'] = "PAGO VECINAL - Historial de Pagos por Propiedad"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:G1')

    # Property information
    ws['A3'] = "Información de la Propiedad"
    ws['A3'].font = Font(size=12, bold=True)
    ws.merge_cells('A3:G3')

    property_info = [
        ["Villa:", property_data['villa']],
        ["Fila:", property_data['row_letter']],
        ["Número:", str(property_data['number'])],
        ["Propietario:", property_data['owner_name']],
        ["Teléfono:", property_data.get('owner_phone', 'N/A')]
    ]

    for i, (label, value) in enumerate(property_info, 5):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)

    # Payments section
    row = 12
    if payments_data:
        ws[f'A{row}'] = "Pagos Realizados"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        ws.merge_cells(f'A{row}:G{row}')

        headers = ["Fecha", "Monto", "Estado", "Referencia"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row+1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        for payment in payments_data:
            row += 1
            ws.cell(row=row, column=1).value = payment['payment_date'].strftime("%d/%m/%Y")
            ws.cell(row=row, column=2).value = float(payment['amount'])
            ws.cell(row=row, column=2).number_format = '"S/ "#,##0.00'
            ws.cell(row=row, column=3).value = payment['status']
            ws.cell(row=row, column=4).value = payment.get('reference', 'N/A')

        # Total payments
        row += 2
        total_payments = sum(p['amount'] for p in payments_data)
        ws[f'A{row}'] = "Total Pagado:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = float(total_payments)
        ws[f'B{row}'].number_format = '"S/ "#,##0.00'
        ws[f'B{row}'].font = Font(bold=True)

    # Outstanding fees section
    if fees_data:
        row += 3
        ws[f'A{row}'] = "Cuotas Pendientes"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        ws.merge_cells(f'A{row}:G{row}')

        headers = ["Período", "Monto", "Fecha Vencimiento", "Estado"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row+1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

        for fee in fees_data:
            row += 1
            ws.cell(row=row, column=1).value = f"{fee['month']}/{fee['year']}"
            ws.cell(row=row, column=2).value = float(fee['amount'])
            ws.cell(row=row, column=2).number_format = '"S/ "#,##0.00'
            ws.cell(row=row, column=3).value = fee['due_date'].strftime("%d/%m/%Y")
            ws.cell(row=row, column=4).value = fee['status']

        # Total outstanding
        row += 2
        total_outstanding = sum(f['amount'] for f in fees_data)
        ws[f'A{row}'] = "Total Pendiente:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = float(total_outstanding)
        ws[f'B{row}'].number_format = '"S/ "#,##0.00'
        ws[f'B{row}'].font = Font(bold=True)

    # Footer
    row += 3
    ws[f'A{row}'] = f"Reporte generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    # Auto-adjust column widths
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_outstanding_fees_excel(fees_data):
    """
    Generate an Excel report of all outstanding fees
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Cuotas Pendientes"

    # Title
    ws['A1'] = "PAGO VECINAL - Cuotas Pendientes - Reporte General"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:F1')

    if fees_data:
        headers = ["Propiedad", "Propietario", "Período", "Monto", "Fecha Vencimiento"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        for row, fee in enumerate(fees_data, 4):
            property_str = f"{fee['property_villa']} - {fee['property_row_letter']}{fee['property_number']}"
            ws.cell(row=row, column=1).value = property_str
            ws.cell(row=row, column=2).value = fee['property_owner_name']
            ws.cell(row=row, column=3).value = f"{fee['month']}/{fee['year']}"
            ws.cell(row=row, column=4).value = float(fee['amount'])
            ws.cell(row=row, column=4).number_format = '"S/ "#,##0.00'
            ws.cell(row=row, column=5).value = fee['due_date'].strftime("%d/%m/%Y")

        # Summary
        row = len(fees_data) + 6
        total_amount = sum(f['amount'] for f in fees_data)
        ws[f'A{row}'] = "Total Cuotas Pendientes:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = len(fees_data)
        ws[f'B{row}'].font = Font(bold=True)

        ws[f'A{row+1}'] = "Monto Total Pendiente:"
        ws[f'A{row+1}'].font = Font(bold=True)
        ws[f'B{row+1}'] = float(total_amount)
        ws[f'B{row+1}'].number_format = '"S/ "#,##0.00'
        ws[f'B{row+1}'].font = Font(bold=True)
    else:
        ws['A3'] = "No hay cuotas pendientes."

    # Footer
    footer_row = len(fees_data) + 9 if fees_data else 5
    ws[f'A{footer_row}'] = f"Reporte generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    # Auto-adjust column widths
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_monthly_payment_summary_excel(year, month, payments_data):
    """
    Generate an Excel summary of payments for a specific month
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Resumen {month:02d}-{year}"

    # Title
    ws['A1'] = f"PAGO VECINAL - Resumen de Pagos - {month:02d}/{year}"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:F1')

    if payments_data:
        headers = ["Propiedad", "Propietario", "Fecha Pago", "Monto", "Estado"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        for row, payment in enumerate(payments_data, 4):
            property_str = f"{payment['property_villa']} - {payment['property_row_letter']}{payment['property_number']}"
            ws.cell(row=row, column=1).value = property_str
            ws.cell(row=row, column=2).value = payment['property_owner_name']
            ws.cell(row=row, column=3).value = payment['payment_date'].strftime("%d/%m/%Y")
            ws.cell(row=row, column=4).value = float(payment['amount'])
            ws.cell(row=row, column=4).number_format = '"S/ "#,##0.00'
            ws.cell(row=row, column=5).value = payment['status']

        # Summary
        row = len(payments_data) + 6
        total_payments = len(payments_data)
        total_amount = sum(p['amount'] for p in payments_data)
        ws[f'A{row}'] = "Total Pagos:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = total_payments
        ws[f'B{row}'].font = Font(bold=True)

        ws[f'A{row+1}'] = "Monto Total Recaudado:"
        ws[f'A{row+1}'].font = Font(bold=True)
        ws[f'B{row+1}'] = float(total_amount)
        ws[f'B{row+1}'].number_format = '"S/ "#,##0.00'
        ws[f'B{row+1}'].font = Font(bold=True)
    else:
        ws['A3'] = "No hay pagos registrados para este período."

    # Footer
    footer_row = len(payments_data) + 9 if payments_data else 5
    ws[f'A{footer_row}'] = f"Reporte generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    # Auto-adjust column widths
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_monthly_fees_excel(fees_data, start_year, start_month, end_year, end_month):
    """
    Generate an Excel report of monthly fees for a specific period range
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Cuotas Mensuales"

    # Title
    ws['A1'] = "PAGO VECINAL - Reporte de Cuotas Mensuales"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:F1')

    # Period information
    period_text = f"Período: {start_month:02d}/{start_year} - {end_month:02d}/{end_year}"
    ws['A3'] = period_text
    ws['A3'].font = Font(size=12, bold=True)
    ws.merge_cells('A3:F3')

    if fees_data:
        # Headers
        headers = ["Propiedad", "Propietario", "Período", "Monto", "Estado", "Fecha Vencimiento"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Data rows
        for row, fee in enumerate(fees_data, 6):
            property_str = f"{fee['property_villa']} - {fee['property_row_letter']}{fee['property_number']}"
            ws.cell(row=row, column=1).value = property_str
            ws.cell(row=row, column=2).value = fee['property_owner_name']
            ws.cell(row=row, column=3).value = f"{fee['month']:02d}/{fee['year']}"
            ws.cell(row=row, column=4).value = float(fee['amount'])
            ws.cell(row=row, column=4).number_format = '"S/ "#,##0.00'
            ws.cell(row=row, column=5).value = fee['status']
            ws.cell(row=row, column=6).value = fee['due_date'].strftime("%d/%m/%Y")

        # Summary section
        summary_row = len(fees_data) + 8
        ws[f'A{summary_row}'] = "Resumen del Reporte"
        ws[f'A{summary_row}'].font = Font(size=12, bold=True)
        ws.merge_cells(f'A{summary_row}:F{summary_row}')

        # Calculate totals by status
        status_counts = {}
        total_amount = 0
        for fee in fees_data:
            status = fee['status']
            if status not in status_counts:
                status_counts[status] = {'count': 0, 'amount': 0}
            status_counts[status]['count'] += 1
            status_counts[status]['amount'] += fee['amount']
            total_amount += fee['amount']

        # Display status summary
        current_row = summary_row + 2
        for status, data in status_counts.items():
            ws[f'A{current_row}'] = f"Cuotas {status.title()}:"
            ws[f'A{current_row}'].font = Font(bold=True)
            ws[f'B{current_row}'] = data['count']
            ws[f'C{current_row}'] = float(data['amount'])
            ws[f'C{current_row}'].number_format = '"S/ "#,##0.00'
            current_row += 1

        # Total summary
        ws[f'A{current_row}'] = "Total Cuotas:"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'] = len(fees_data)
        ws[f'B{current_row}'].font = Font(bold=True)

        ws[f'A{current_row+1}'] = "Monto Total:"
        ws[f'A{current_row+1}'].font = Font(bold=True)
        ws[f'B{current_row+1}'] = float(total_amount)
        ws[f'B{current_row+1}'].number_format = '"S/ "#,##0.00'
        ws[f'B{current_row+1}'].font = Font(bold=True)
    else:
        ws['A5'] = "No hay cuotas registradas para el período especificado."

    # Footer
    footer_row = len(fees_data) + 12 if fees_data else 7
    ws[f'A{footer_row}'] = f"Reporte generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    # Auto-adjust column widths
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_annual_property_statement_excel(property_data, year, fees_data, payments_data):
    """
    Generate an annual Excel statement for a property
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Estado Anual {year}"

    # Title
    ws['A1'] = "PAGO VECINAL - Estado Anual por Propiedad"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:G1')

    # Property information
    ws['A3'] = "Información de la Propiedad"
    ws['A3'].font = Font(size=12, bold=True)
    ws.merge_cells('A3:G3')

    property_info = [
        ["Villa:", property_data['villa']],
        ["Fila:", property_data['row_letter']],
        ["Número:", str(property_data['number'])],
        ["Propietario:", property_data['owner_name']],
        ["Teléfono:", property_data.get('owner_phone', 'N/A')]
    ]

    for i, (label, value) in enumerate(property_info, 5):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)

    # Annual summary
    row = 12
    total_fees = sum(f['amount'] for f in fees_data) if fees_data else 0
    total_payments = sum(p['amount'] for p in payments_data) if payments_data else 0
    balance = total_fees - total_payments

    summary_data = [
        ["Total Cuotas Generadas:", f"S/ {total_fees:.2f}"],
        ["Total Pagos Realizados:", f"S/ {total_payments:.2f}"],
        ["Saldo Pendiente:", f"S/ {balance:.2f}"]
    ]

    for i, (label, value) in enumerate(summary_data, row):
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = Font(bold=True)
        if "S/" in value:
            ws[f'B{i}'] = float(value.replace("S/ ", ""))
            ws[f'B{i}'].number_format = '"S/ "#,##0.00'
        else:
            ws[f'B{i}'] = value
        ws[f'B{i}'].font = Font(bold=True)

    # Monthly breakdown
    row += 6
    ws[f'A{row}'] = "Detalle por Mes"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    ws.merge_cells(f'A{row}:G{row}')

    monthly_headers = ["Mes", "Cuotas", "Pagos", "Saldo"]
    for col, header in enumerate(monthly_headers, 1):
        cell = ws.cell(row=row+1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Create monthly breakdown
    monthly_data = {}
    for month in range(1, 13):
        monthly_data[month] = {'fees': 0, 'payments': 0}

    for fee in fees_data:
        if fee['month'] in monthly_data:
            monthly_data[fee['month']]['fees'] += fee['amount']

    for payment in payments_data:
        payment_month = payment['payment_date'].month
        if payment_month in monthly_data:
            monthly_data[payment_month]['payments'] += payment['amount']

    month_names = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                   "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

    for month in range(1, 13):
        row += 1
        fees = monthly_data[month]['fees']
        payments = monthly_data[month]['payments']
        balance = fees - payments
        ws.cell(row=row, column=1).value = month_names[month]
        ws.cell(row=row, column=2).value = float(fees)
        ws.cell(row=row, column=2).number_format = '"S/ "#,##0.00'
        ws.cell(row=row, column=3).value = float(payments)
        ws.cell(row=row, column=3).number_format = '"S/ "#,##0.00'
        ws.cell(row=row, column=4).value = float(balance)
        ws.cell(row=row, column=4).number_format = '"S/ "#,##0.00'

    # Footer
    row += 3
    ws[f'A{row}'] = f"Estado generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    # Auto-adjust column widths
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 15

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer